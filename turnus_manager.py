"""
Turnusplan-manager - håndterer Excel-fila for vaktplanen.
Oppretter, leser og oppdaterer turnusplanen med fargekoding.
"""

import os
import platform
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

EXCEL_FILE = "turnus.xlsx"

# Farger
RED_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
GREEN_FILL = PatternFill(start_color="44CC44", end_color="44CC44", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFDD44", end_color="FFDD44", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
DARK_FONT = Font(color="000000", bold=False, size=11)
BOLD_FONT = Font(color="000000", bold=True, size=11)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def get_dates_for_week():
    """Hjelpefunksjon for å finne datoer for gjeldende uke (mandag-fredag)"""
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    dates = []
    for i in range(5):  # Mandag til fredag
        day = monday + timedelta(days=i)
        dates.append(day.strftime("%d.%m.%Y"))
    return dates

def create_turnus_excel():
    """
    Oppretter en turnusplan med fiktive vakter for en uke.
    Returnerer stien til fila.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Turnusplan"

    # Bredde på kolonner
    ws.column_dimensions['A'].width = 22
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

    dates = get_dates_for_week()

    # Header-rad
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "🏥 VESTRE VIKEN - TURNUSPLAN"
    cell.font = Font(size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35

    # Kolonneoverskrifter
    headers = ["Navn"] + [f"{day}\n{date}" for day, date in zip(["Mandag", "Tirsdag", "Onsdag", "Torsdag", "Fredag"], dates)]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[2].height = 40

    # Turnusdata (fiktivt oppsett)
    turnus_data = [
        ["Nils Dagenderpå", "DAG 🕐 08-20", "NATT 🌙 20-08", "FRI", "DAG 🕐 08-20", "DAG 🕐 08-20"],
        ["Mette Prada Hansen", "KVLD 🌆 14-22", "FRI", "DAG 🕐 08-20", "FRI", "KVLD 🌆 14-22"],
        ["Wes Side Story", "NATT 🌙 20-08", "DAG 🕐 08-20", "FRI", "NATT 🌙 20-08", "FRI"],
        ["Dr. Anton Graff", "FRI", "KVLD 🌆 14-22", "KVLD 🌆 14-22", "FRI", "NATT 🌙 20-08"],
        ["Kari Vaktmester", "DAG 🕐 08-20", "DAG 🕐 08-20", "NATT 🌙 20-08", "FRI", "FRI"],
        ["Ole Tidsklemme", "FRI", "FRI", "DAG 🕐 08-20", "DAG 🕐 08-20", "KVLD 🌆 14-22"],
        ["Lise Trøtt", "NATT 🌙 20-08", "NATT 🌙 20-08", "FRI", "KVLD 🌆 14-22", "FRI"],
    ]

    for row_idx, row_data in enumerate(turnus_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DARK_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            if col_idx > 1:
                # Fargekoding basert på vakttype
                if "DAG" in str(value):
                    cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                elif "NATT" in str(value):
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                elif "KVLD" in str(value):
                    cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                elif "FRI" in str(value):
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        ws.row_dimensions[row_idx].height = 25

    # Legg til en "Forklaring"-rad
    forklaring_row = len(turnus_data) + 4
    ws.merge_cells(f'A{forklaring_row}:F{forklaring_row}')
    cell = ws.cell(row=forklaring_row, column=1, value="📋 Fargekodeforklaring")
    cell.font = Font(bold=True, size=11)

    forklaringer = [
        (forklaring_row+1, "🕐 DAG 08-20 (grønn)", "E8F5E9"),
        (forklaring_row+2, "🌙 NATT 20-08 (blå)", "E3F2FD"),
        (forklaring_row+3, "🌆 KVELD 14-22 (oransje)", "FFF3E0"),
        (forklaring_row+4, "❌ SYK (rød)", "FF4444"),
        (forklaring_row+5, "✅ VIKAR (grønn)", "44CC44"),
    ]
    for row, text, color in forklaringer:
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.border = thin_border

    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), EXCEL_FILE)
    wb.save(path)
    return path


def load_or_create_turnus():
    """Laster turnusplanen eller oppretter en ny hvis den ikke finnes."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), EXCEL_FILE)
    
    if not os.path.exists(path):
        print("📊 Oppretter ny turnusplan...")
        path = create_turnus_excel()
    
    return path


def mark_sick(path, staff_name):
    """
    Markerer en ansatt som syk i turnusplanen.
    Finner dagens vakt og setter rød bakgrunn.
    """
    wb = load_workbook(path)
    ws = wb.active

    today_col = datetime.now().weekday() + 2  # Kolonne B=2 (mandag) til F=6 (fredag)
    if today_col > 6:
        today_col = 2  # Hvis helg, marker mandag

    for row in range(3, 10):
        cell = ws.cell(row=row, column=1)
        if cell.value and staff_name.lower() in str(cell.value).lower():
            shift_cell = ws.cell(row=row, column=today_col)
            original_value = shift_cell.value
            shift_cell.value = f"❌ SYK ({original_value})"
            shift_cell.fill = RED_FILL
            shift_cell.font = Font(color="FFFFFF", bold=True, size=11)
            print(f"  ☑️  Markert {staff_name} som syk på dagens vakt")
            break

    wb.save(path)
    return path


def mark_replacement(path, sick_name, replacement_name):
    """
    Markerer at en vikar har tatt vakten.
    """
    wb = load_workbook(path)
    ws = wb.active

    today_col = datetime.now().weekday() + 2
    if today_col > 6:
        today_col = 2

    for row in range(3, 10):
        cell = ws.cell(row=row, column=1)
        if cell.value and replacement_name.lower() in str(cell.value).lower():
            shift_cell = ws.cell(row=row, column=today_col)
            shift_cell.value = f"✅ VIKAR for {sick_name}"
            shift_cell.fill = GREEN_FILL
            shift_cell.font = Font(color="FFFFFF", bold=True, size=11)
            print(f"  ☑️  Markert {replacement_name} som vikar for {sick_name}")
            break

    wb.save(path)
    return path


def open_excel(path):
    """Åpner Excel-fila i standardprogrammet."""
    try:
        system = platform.system()
        if system == 'Darwin':
            os.system(f"open \"{path}\"")
        elif system == 'Windows':
            os.startfile(path)
        else:
            os.system(f"xdg-open \"{path}\"")
        print("📊 Turnusplan åpnet!")
    except Exception as e:
        print(f"⚠️  Kunne ikke åpne Excel automatisk: {e}")
        print(f"📁 Åpne filen manuelt: {path}")


if __name__ == "__main__":
    # Test - opprett turnusplan
    path = load_or_create_turnus()
    print(f"Turnusplan opprettet: {path}")
    open_excel(path)
