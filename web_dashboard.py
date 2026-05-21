"""
Web Dashboard for live-visning av turnusplanen.
Leser direkte fra Excel-filen og viser den som HTML.
"""

from flask import render_template_string
from openpyxl import load_workbook
import os
from turnus_manager import EXCEL_FILE, get_dates_for_week

HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Live Turnusplan - Vestre Viken</title>
    <meta http-equiv="refresh" content="2"> <!-- Auto-oppdaterer hvert 2. sekund -->
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
            background-color: #f5f7fa;
            margin: 0;
            padding: 40px;
            color: #333;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.1);
            overflow: hidden;
        }
        .header {
            background-color: #1F4E79;
            color: white;
            padding: 25px;
            text-align: center;
        }
        .header h1 {
            margin: 0;
            font-size: 28px;
            letter-spacing: 1px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
        }
        th {
            background-color: #2F5496;
            color: white;
            padding: 15px;
            text-align: center;
            font-weight: 600;
            font-size: 16px;
            border: 1px solid #1F4E79;
            white-space: pre-line;
        }
        td {
            padding: 15px;
            text-align: center;
            border: 1px solid #e1e4e8;
            font-size: 15px;
            font-weight: 500;
        }
        td:first-child {
            text-align: left;
            font-weight: bold;
            background-color: #f8f9fa;
        }
        
        /* Fargekoder for vakter */
        .vakt-dag { background-color: #E8F5E9; color: #2E7D32; }
        .vakt-natt { background-color: #E3F2FD; color: #1565C0; }
        .vakt-kveld { background-color: #FFF3E0; color: #E65100; }
        .vakt-fri { background-color: #F5F5F5; color: #757575; }
        
        /* MAGIEN: Live endringer */
        .vakt-syk { 
            background-color: #FFEBEE !important; 
            color: #C62828 !important;
            font-weight: bold;
            animation: pulse-red 2s infinite;
        }
        .vakt-vikar { 
            background-color: #E8F5E9 !important; 
            color: #2E7D32 !important;
            font-weight: bold;
            border: 2px solid #4CAF50 !important;
            animation: pulse-green 2s infinite;
        }

        @keyframes pulse-red {
            0% { box-shadow: 0 0 0 0 rgba(244, 67, 54, 0.4); }
            70% { box-shadow: 0 0 0 10px rgba(244, 67, 54, 0); }
            100% { box-shadow: 0 0 0 0 rgba(244, 67, 54, 0); }
        }
        @keyframes pulse-green {
            0% { box-shadow: 0 0 0 0 rgba(76, 175, 80, 0.4); }
            70% { box-shadow: 0 0 0 10px rgba(76, 175, 80, 0); }
            100% { box-shadow: 0 0 0 0 rgba(76, 175, 80, 0); }
        }

        .legend {
            display: flex;
            justify-content: center;
            gap: 20px;
            padding: 20px;
            background: #f8f9fa;
            border-top: 2px solid #e1e4e8;
        }
        .legend-item {
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 14px;
            font-weight: bold;
        }
        .color-box {
            width: 20px;
            height: 20px;
            border-radius: 4px;
            border: 1px solid #ccc;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🏥 VESTRE VIKEN - LIVE TURNUSPLAN</h1>
        </div>
        <table>
            <tr>
                <th>Navn</th>
                {% for header in headers %}
                <th>{{ header }}</th>
                {% endfor %}
            </tr>
            {% for row in rows %}
            <tr>
                <td>{{ row.name }}</td>
                {% for cell in row.cells %}
                <td class="{{ cell.css_class }}">{{ cell.value }}</td>
                {% endfor %}
            </tr>
            {% endfor %}
        </table>
        <div class="legend">
            <div class="legend-item"><div class="color-box vakt-dag"></div> DAG 08-20</div>
            <div class="legend-item"><div class="color-box vakt-natt"></div> NATT 20-08</div>
            <div class="legend-item"><div class="color-box vakt-kveld"></div> KVELD 14-22</div>
            <div class="legend-item"><div class="color-box vakt-syk" style="animation:none;"></div> SYK</div>
            <div class="legend-item"><div class="color-box vakt-vikar" style="animation:none;"></div> VIKAR</div>
        </div>
    </div>
</body>
</html>
"""

def get_css_class(val):
    """Finner riktig CSS-klasse basert på celleteksten."""
    val_str = str(val).upper() if val else ""
    if "SYK" in val_str:
        return "vakt-syk"
    if "VIKAR" in val_str:
        return "vakt-vikar"
    if "DAG" in val_str:
        return "vakt-dag"
    if "NATT" in val_str:
        return "vakt-natt"
    if "KVLD" in val_str or "KVELD" in val_str:
        return "vakt-kveld"
    if "FRI" in val_str:
        return "vakt-fri"
    return ""

from turnus_manager import load_or_create_turnus

def render_live_dashboard():
    """Leser Excel-fila og bygger HTML."""
    path = load_or_create_turnus()
        
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    
    # Hent headers (rad 2, kolonne B til F)
    dates = get_dates_for_week()
    days = ["Mandag", "Tirsdag", "Onsdag", "Torsdag", "Fredag"]
    headers = [f"{day}\\n{date}" for day, date in zip(days, dates)]
    
    # Hent rad-data (fra rad 3 til vi treffer tom eller forklaring)
    rows = []
    for row_idx in range(3, 15):
        name_cell = ws.cell(row=row_idx, column=1).value
        if not name_cell or str(name_cell).startswith("📋"):
            break
            
        cells = []
        for col_idx in range(2, 7): # Kolonne B(2) til F(6)
            val = ws.cell(row=row_idx, column=col_idx).value or ""
            cells.append({
                "value": val,
                "css_class": get_css_class(val)
            })
            
        rows.append({
            "name": name_cell,
            "cells": cells
        })
        
    return render_template_string(HTML_TEMPLATE, headers=headers, rows=rows)
